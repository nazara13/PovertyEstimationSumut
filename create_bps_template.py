"""
Membuat template Excel BPS untuk diisi data SUSENAS.
Jalankan: python create_bps_template.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

kabupaten_sumut = [
    ('1201', 'Kabupaten Nias'),
    ('1202', 'Kabupaten Mandailing Natal'),
    ('1203', 'Kabupaten Tapanuli Selatan'),
    ('1204', 'Kabupaten Tapanuli Tengah'),
    ('1205', 'Kabupaten Tapanuli Utara'),
    ('1206', 'Kabupaten Toba'),
    ('1207', 'Kabupaten Labuhanbatu'),
    ('1208', 'Kabupaten Asahan'),
    ('1209', 'Kabupaten Simalungun'),
    ('1210', 'Kabupaten Dairi'),
    ('1211', 'Kabupaten Karo'),
    ('1212', 'Kabupaten Deli Serdang'),
    ('1213', 'Kabupaten Langkat'),
    ('1214', 'Kabupaten Nias Selatan'),
    ('1215', 'Kabupaten Humbang Hasundutan'),
    ('1216', 'Kabupaten Pakpak Bharat'),
    ('1217', 'Kabupaten Samosir'),
    ('1218', 'Kabupaten Serdang Bedagai'),
    ('1219', 'Kabupaten Batu Bara'),
    ('1220', 'Kabupaten Padang Lawas Utara'),
    ('1221', 'Kabupaten Padang Lawas'),
    ('1222', 'Kabupaten Labuhanbatu Selatan'),
    ('1223', 'Kabupaten Labuhanbatu Utara'),
    ('1224', 'Kabupaten Nias Utara'),
    ('1225', 'Kabupaten Nias Barat'),
    ('1271', 'Kota Sibolga'),
    ('1272', 'Kota Tanjungbalai'),
    ('1273', 'Kota Pematangsiantar'),
    ('1274', 'Kota Tebing Tinggi'),
    ('1275', 'Kota Medan'),
    ('1276', 'Kota Binjai'),
    ('1277', 'Kota Padangsidimpuan'),
    ('1278', 'Kota Gunungsitoli'),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Data Kemiskinan Sumut 2023'

# Style header
header_fill  = PatternFill('solid', fgColor='1F3864')
header_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border  = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)

headers = [
    ('kode_kab',         'Kode BPS Kabupaten', 15),
    ('kabupaten',        'Nama Kabupaten/Kota', 30),
    ('poverty_rate',     'Persentase Penduduk Miskin P0 (%)', 22),
    ('poverty_depth',    'Indeks Kedalaman Kemiskinan P1', 22),
    ('poverty_severity', 'Indeks Keparahan Kemiskinan P2', 22),
    ('garis_kemiskinan', 'Garis Kemiskinan (Rp/kap/bln)', 24),
    ('jumlah_miskin',    'Jumlah Penduduk Miskin (ribu jiwa)', 26),
    ('tahun',            'Tahun Survei', 14),
]

for col_idx, (_, col_name, col_width) in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.fill   = header_fill
    cell.font   = header_font
    cell.alignment = header_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = col_width

ws.row_dimensions[1].height = 40

# Isi data kabupaten (kolom kode & nama saja, sisanya diisi user)
alt_fill = PatternFill('solid', fgColor='F2F2F2')
for row_idx, (kode, nama) in enumerate(kabupaten_sumut, start=2):
    ws.cell(row=row_idx, column=1, value=kode)
    ws.cell(row=row_idx, column=2, value=nama)
    ws.cell(row=row_idx, column=8, value=2023)  # Tahun default
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if row_idx % 2 == 0:
            cell.fill = alt_fill

# Freeze header row
ws.freeze_panes = 'A2'

# Instruksi di sheet kedua
ws2 = wb.create_sheet('Petunjuk')
ws2['A1'] = '📌 PETUNJUK PENGISIAN DATA'
ws2['A1'].font = Font(bold=True, size=13, color='1F3864')
instructions = [
    ('poverty_rate',     'Persentase (%) penduduk miskin. Sumber: Tabel BPS "Persentase Penduduk Miskin".'),
    ('poverty_depth',    'Indeks P1 (Poverty Gap Index). Mengukur rata-rata jarak antara pengeluaran penduduk miskin dan garis kemiskinan.'),
    ('poverty_severity', 'Indeks P2 (Poverty Severity Index). Mengukur ketimpangan pengeluaran antar penduduk miskin.'),
    ('garis_kemiskinan', 'Garis kemiskinan dalam Rupiah per kapita per bulan.'),
    ('jumlah_miskin',    'Jumlah penduduk miskin dalam satuan ribu jiwa.'),
]
for i, (field, desc) in enumerate(instructions, start=3):
    ws2[f'A{i}'] = field
    ws2[f'B{i}'] = desc
    ws2[f'A{i}'].font = Font(bold=True, color='1F3864')
ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 80

out_path = os.path.join('data', 'raw', 'bps_kemiskinan.xlsx')
os.makedirs(os.path.dirname(out_path), exist_ok=True)
wb.save(out_path)
print(f"✅ Template Excel BPS berhasil dibuat: {out_path}")
print(f"   Silakan isi data kemiskinan dari situs BPS: https://sumut.bps.go.id")
